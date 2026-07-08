"""Génération de fichiers Excel : WBS, Gantt, Frais de déplacement."""
import io
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from models.schemas import WizardState, EstimationResult

# ── Couleurs ──────────────────────────────────────────────────────────────────
C_DARK   = "1F4E79"
C_MID    = "2E86C1"
C_LIGHT  = "D6E4F0"
C_XL     = "EBF5FB"
C_WHITE  = "FFFFFF"
C_TOTAL  = "154360"
C_GREEN  = "1E8449"
C_GOLD   = "D4AC0D"

PHASE_FILLS = {
    'Pré-cadrage':        "5DADE2",
    'Cadrage':            "2E86C1",
    'Développement':      "1F4E79",
    'Tests':              "148F77",
    'Mise en Production': "1E8449",
    'Support':            "D4AC0D",
}
PHASE_ORDER = ['Pré-cadrage', 'Cadrage', 'Développement', 'Tests', 'Mise en Production', 'Support']


def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _hf(color=C_WHITE, bold=True, size=10):
    return Font(bold=bold, color=color, size=size)


def _fill(color: str):
    return PatternFill("solid", fgColor=color)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


class ExcelGenerator:
    def generate(
        self,
        project_name: str,
        wizard_state: WizardState,
        estimation_result: EstimationResult,
    ) -> bytes:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        self._sheet_summary(wb, project_name, wizard_state, estimation_result)
        self._sheet_wbs(wb, estimation_result, wizard_state)
        self._sheet_gantt(wb, estimation_result)
        self._sheet_travel(wb, wizard_state, estimation_result)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── FEUILLE 1 : Résumé ────────────────────────────────────────────────────
    def _sheet_summary(self, wb, project_name, wizard_state, estimation_result):
        ws = wb.create_sheet("Résumé")
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 28

        info = wizard_state.project_info
        cur  = info.currency if info else ""

        # Titre
        ws.merge_cells("A1:B1")
        c = ws["A1"]
        c.value = f"📊 {project_name}"
        c.font  = _hf(size=14)
        c.fill  = _fill(C_DARK)
        c.alignment = _center()
        ws.row_dimensions[1].height = 36

        # Infos projet
        info_rows = [
            ("Client",          info.client_name if info else "-"),
            ("Région",          info.region if info else "-"),
            ("Devise",          info.currency if info else "-"),
            ("Date de début",   info.start_date if info else "-"),
            ("Date du document",info.document_date if info else "-"),
            ("Préparé par",     info.prepared_by if info else "-"),
            ("Confidentialité", info.confidentiality if info else "-"),
        ]
        for i, (label, val) in enumerate(info_rows, 2):
            lc = ws.cell(row=i, column=1, value=label)
            vc = ws.cell(row=i, column=2, value=val)
            lc.font = _hf(C_DARK); lc.fill = _fill(C_LIGHT)
            lc.border = vc.border = _border()
            vc.fill = _fill(C_XL) if i % 2 == 0 else _fill(C_WHITE)

        # KPIs
        start = len(info_rows) + 3
        ws.merge_cells(f"A{start}:B{start}")
        kh = ws.cell(row=start, column=1, value="Indicateurs Clés")
        kh.font = _hf(); kh.fill = _fill(C_DARK); kh.alignment = _center()
        ws.row_dimensions[start].height = 22

        kpis = [
            ("Total JH",               f"{estimation_result.total_jh:.1f}"),
            ("Coût Total",             f"{estimation_result.total_cost:,.0f} {cur}"),
            (f"Marge ({wizard_state.margin_percentage:.0f}%)",
                                       f"{estimation_result.margin:,.0f} {cur}"),
            ("Prix de Vente",          f"{estimation_result.selling_price:,.0f} {cur}"),
            ("Frais de Déplacement",   f"{estimation_result.travel_cost:,.0f} {cur}"),
            ("PRIX TOTAL",             f"{estimation_result.total_price:,.0f} {cur}"),
        ]
        for j, (label, val) in enumerate(kpis, start + 1):
            lc = ws.cell(row=j, column=1, value=label)
            vc = ws.cell(row=j, column=2, value=val)
            lc.border = vc.border = _border()
            vc.alignment = Alignment(horizontal="right")
            if label == "PRIX TOTAL":
                lc.font = vc.font = _hf()
                lc.fill = vc.fill = _fill(C_TOTAL)
            else:
                lc.font = _hf(C_DARK)
                lc.fill = _fill(C_LIGHT) if j % 2 == 0 else _fill(C_WHITE)

    # ── FEUILLE 2 : Planning WBS (AVEC HIÉRARCHIE) ──────────────────────────
    def _sheet_wbs(self, wb, estimation_result, wizard_state):
        ws = wb.create_sheet("Planning WBS")
        cal = (estimation_result.planning and
               estimation_result.planning.config.calendar == "islamic")

        # ✅ En-têtes avec Niveau pour hiérarchie
        if cal:
            headers = ["Niv", "ID", "Segment", "Module", "Phase", "Tâche", "Durée (j)",
                       "Début (G)", "Début (H)", "Fin (G)", "Fin (H)",
                       "Sur site", "À dist.", "JH Total", "Dépendances"]
            col_widths = [4, 6, 14, 18, 18, 30, 10, 13, 15, 13, 15, 8, 8, 10, 16]
        else:
            headers = ["Niv", "ID", "Segment", "Module", "Phase", "Tâche", "Durée (j)",
                       "Date Début", "Date Fin", "Sur site", "À dist.", "JH Total", "Dépendances"]
            col_widths = [4, 6, 14, 18, 18, 30, 10, 13, 13, 8, 8, 10, 16]

        for i, (h, w) in enumerate(zip(headers, col_widths), 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = _hf(); c.fill = _fill(C_DARK)
            c.alignment = _center(); c.border = _border()
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        if not estimation_result.planning:
            ws.cell(row=2, column=1, value="Aucun planning calculé.")
            return

        row_idx = 2
        for task in estimation_result.planning.tasks:
            ph_fill = _fill(PHASE_FILLS.get(task.phase, C_MID))
            alt = _fill(C_XL) if row_idx % 2 == 0 else _fill(C_WHITE)
            
            # ✅ Indentation selon le niveau hiérarchique
            level = getattr(task, 'level', 0)
            indent = "  " * level
            task_name = f"{indent}{task.name}"
            
            # ✅ Si c'est une phase (niveau 0), mettre en gras
            is_phase = getattr(task, 'is_phase', False)
            
            if cal:
                vals = [
                    level,
                    task.id,
                    task.segment,
                    task.module,
                    task.phase,
                    task_name,
                    task.duration_days,
                    task.start_date,
                    task.start_date_hijri or "-",
                    task.end_date,
                    task.end_date_hijri or "-",
                    task.onsite_resources,
                    task.remote_resources,
                    task.total_jh,
                    ", ".join(task.dependencies)
                ]
            else:
                vals = [
                    level,
                    task.id,
                    task.segment,
                    task.module,
                    task.phase,
                    task_name,
                    task.duration_days,
                    task.start_date,
                    task.end_date,
                    task.onsite_resources,
                    task.remote_resources,
                    task.total_jh,
                    ", ".join(task.dependencies)
                ]

            for col_idx, val in enumerate(vals, 1):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.border = _border()
                
                # Colonne Phase = couleur de la phase
                if col_idx == 5:  # Phase
                    c.fill = ph_fill
                    c.font = Font(bold=True, color=C_WHITE, size=9)
                else:
                    c.fill = alt
                    # ✅ Si c'est une phase (niveau 0), mettre en gras
                    if is_phase and col_idx in [6]:  # Colonne Tâche
                        c.font = Font(bold=True, size=10)
                
                c.alignment = _left()

            row_idx += 1

    # ── FEUILLE 3 : Gantt ─────────────────────────────────────────────────────
    def _sheet_gantt(self, wb, estimation_result):
        ws = wb.create_sheet("Gantt")
        planning = estimation_result.planning

        if not planning or not planning.tasks:
            ws.cell(row=1, column=1, value="Aucun planning calculé.")
            return

        # Récupérer les bornes de dates
        try:
            all_starts = [datetime.strptime(t.start_date, "%Y-%m-%d") for t in planning.tasks]
            all_ends   = [datetime.strptime(t.end_date,   "%Y-%m-%d") for t in planning.tasks]
            proj_start = min(all_starts)
            proj_end   = max(all_ends)
        except Exception:
            ws.cell(row=1, column=1, value="Erreur de calcul des dates.")
            return

        total_days = max(1, (proj_end - proj_start).days)
        num_weeks = total_days // 7 + 1

        # En-têtes fixes
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 16

        header_row = ["ID", "Phase", "Tâche", "JH"]
        for i, h in enumerate(header_row, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = _hf(); c.fill = _fill(C_DARK)
            c.alignment = _center(); c.border = _border()

        # En-têtes de semaines
        col_offset = 5
        for w in range(num_weeks):
            week_start = proj_start + timedelta(weeks=w)
            c = ws.cell(row=1, column=col_offset + w,
                        value=week_start.strftime("%d/%m"))
            c.font = _hf(size=8); c.fill = _fill(C_MID)
            c.alignment = _center()
            ws.column_dimensions[get_column_letter(col_offset + w)].width = 4

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "E2"

        # Lignes de tâches
        for row_idx, task in enumerate(planning.tasks, 2):
            ph_fill_hex = PHASE_FILLS.get(task.phase, C_MID)
            is_phase = getattr(task, 'is_phase', False)

            for col_idx, val in enumerate(
                [task.id, task.phase, task.name, task.total_jh], 1
            ):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.border = _border()
                c.fill = _fill(ph_fill_hex) if col_idx == 2 else (
                    _fill(C_XL) if row_idx % 2 == 0 else _fill(C_WHITE)
                )
                if col_idx == 2:
                    c.font = Font(bold=True, color=C_WHITE, size=9)
                elif col_idx == 3 and is_phase:
                    c.font = Font(bold=True, size=10)
                c.alignment = _left()

            # Barres de Gantt
            try:
                t_start = datetime.strptime(task.start_date, "%Y-%m-%d")
                t_end   = datetime.strptime(task.end_date,   "%Y-%m-%d")
                start_w = (t_start - proj_start).days // 7
                end_w   = (t_end   - proj_start).days // 7
                for w in range(start_w, min(end_w + 1, num_weeks)):
                    c = ws.cell(row=row_idx, column=col_offset + w, value="")
                    c.fill = _fill(ph_fill_hex)
            except Exception:
                pass

    # ── FEUILLE 4 : Frais de déplacement ─────────────────────────────────────
    def _sheet_travel(self, wb, wizard_state, estimation_result):
        ws = wb.create_sheet("Frais Déplacement")
        travel = wizard_state.travel_costs
        cur = wizard_state.project_info.currency if wizard_state.project_info else ""

        if not travel or not travel.expenses:
            ws.cell(row=1, column=1, value="Aucun frais de déplacement renseigné.")
            return

        # Paramètres
        params = [
            ("Per diem / jour", f"{travel.per_diem} {cur}"),
            ("Hôtel / nuit",    f"{travel.hotel_rate} {cur}"),
            ("Vol A/R",         f"{travel.flight_cost} {cur}"),
        ]
        for i, (k, v) in enumerate(params, 1):
            lc = ws.cell(row=i, column=1, value=k)
            vc = ws.cell(row=i, column=2, value=v)
            lc.font = _hf(C_DARK); lc.fill = _fill(C_LIGHT)
            lc.border = vc.border = _border()

        # Tableau détaillé
        headers    = ["Phase","Nbre Ressources","Jours","Per Diem","Hôtel","Vol","Total Phase"]
        col_widths = [18, 16, 8, 14, 14, 14, 18]
        start_row  = len(params) + 2

        for i, (h, w) in enumerate(zip(headers, col_widths), 1):
            c = ws.cell(row=start_row, column=i, value=h)
            c.font = _hf(); c.fill = _fill(C_DARK)
            c.alignment = _center(); c.border = _border()
            ws.column_dimensions[get_column_letter(i)].width = w

        grand_total = 0.0
        for row_idx, exp in enumerate(travel.expenses, start_row + 1):
            try:
                s = datetime.strptime(exp.start_date, "%Y-%m-%d")
                e = datetime.strptime(exp.end_date,   "%Y-%m-%d")
                days = max(1, (e - s).days)
            except Exception:
                days = 1
            n        = exp.onsite_resources
            per_diem = days * travel.per_diem * n
            hotel    = days * travel.hotel_rate * n
            flight   = travel.flight_cost * n
            total_ph = per_diem + hotel + flight
            grand_total += total_ph

            vals = [exp.phase, n, days,
                    f"{per_diem:,.0f}", f"{hotel:,.0f}", f"{flight:,.0f}",
                    f"{total_ph:,.0f} {cur}"]
            alt = _fill(C_XL) if row_idx % 2 == 0 else _fill(C_WHITE)
            for col_idx, val in enumerate(vals, 1):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.fill = alt; c.border = _border()

        # Ligne total
        tr = start_row + len(travel.expenses) + 1
        ws.cell(row=tr, column=1, value="TOTAL").font = _hf()
        ws.cell(row=tr, column=1).fill = _fill(C_TOTAL)
        tc = ws.cell(row=tr, column=7, value=f"{grand_total:,.0f} {cur}")
        tc.font = _hf(); tc.fill = _fill(C_TOTAL)
        for col in range(1, 8):
            ws.cell(row=tr, column=col).border = _border()